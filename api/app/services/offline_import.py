from __future__ import annotations

import csv
import fnmatch
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.services.normalize_keys import normalize_registration_no
from app.services.registration_no_parser import parse_registration_no


DEFAULT_PATTERN = "*.csv,*.xlsx,*.xls,*.json,*.ndjson"
DEFAULT_SOURCE_KEY = "nmpa_legacy_dump"
DEFAULT_IMPORT_ROOT = Path("/data/import")
REGNO_HINT_KEYS = (
    "注册证号",
    "注册证编号",
    "备案号",
    "备案编号",
    "证号",
    "注册编号",
    "注册号",
    "许可证号",
    "批准文号",
)
PRODUCTION_ADDR_KEYS = ("生产地址", "生产场所", "生产企业住所", "生产企业地址")
REGISTRANT_ADDR_KEYS = ("注册人住所", "注册人地址", "注册地址")
COUNTRY_PARAM_EXTRACT_VERSION = "country_region_v1_20260220"
ORIGIN_PARAM_EXTRACT_VERSION = "origin_bucket_v1_20260220"

_REGION_PATTERNS: list[tuple[str, str]] = [
    (r"(香港|香港特别行政区|Hong\s*Kong)", "中国香港"),
    (r"(澳门|澳門|澳门特别行政区|Macao|Macau)", "中国澳门"),
    (r"(台湾|台灣|Taiwan)", "中国台湾"),
]

_COUNTRY_PATTERNS: list[tuple[str, str]] = [
    (r"(中国|China\b)", "中国"),
    (r"(美国|U\.?S\.?A\.?\b|United\s+States\b)", "美国"),
    (r"(德国|Germany\b)", "德国"),
    (r"(日本|Japan\b)", "日本"),
    (r"(法国|France\b)", "法国"),
    (r"(英国|U\.?K\.?\b|United\s+Kingdom\b|Britain\b)", "英国"),
    (r"(比利时|Belgium\b)", "比利时"),
    (r"(意大利|Italy\b)", "意大利"),
    (r"(瑞士|Switzerland\b)", "瑞士"),
    (r"(韩国|South\s*Korea\b|Korea\b)", "韩国"),
    (r"(新加坡|Singapore\b)", "新加坡"),
    (r"(爱尔兰|Ireland\b)", "爱尔兰"),
    (r"(加拿大|Canada\b)", "加拿大"),
    (r"(荷兰|Netherlands\b|Holland\b)", "荷兰"),
    (r"(西班牙|Spain\b)", "西班牙"),
    (r"(葡萄牙|Portugal\b)", "葡萄牙"),
    (r"(印度|India\b)", "印度"),
    (r"(波多黎各|Puerto\s*Rico\b|\bPR\b)", "波多黎各"),
]


@dataclass
class FileImportResult:
    storage_uri: str
    relative_path: str
    file_sha256: str
    file_size: int
    file_mtime: datetime
    imported: bool
    skipped_reason: str | None
    rows_scanned: int
    rows_written: int
    rows_failed: int


@dataclass
class OfflineImportResult:
    source_key: str
    dataset_version: str
    dataset_id: str
    root_path: str
    recursive: bool
    max_depth: int
    pattern: str
    only_new: bool
    dry_run: bool
    files_scanned: int
    files_imported: int
    files_skipped: int
    rows_written: int
    rows_failed: int
    new_files_count: int
    dup_files_count: int
    ext_filtered_count: int
    parse_level_distribution: dict[str, int]
    top_parse_reasons: list[dict[str, int | str]]
    action_suffix_counts: dict[str, int]
    issuer_alias_counts: dict[str, int]
    country_region_counts: dict[str, int]
    origin_bucket_counts: dict[str, int]
    product_params_written: int


@dataclass(frozen=True)
class CountryExtractionResult:
    country_or_region: str
    geo_type: str  # country | region
    source_field: str
    raw_value: str
    confidence: float


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _dataset_version() -> str:
    return _now().strftime("%Y%m%d_%H%M%S")


def _parse_patterns(pattern: str) -> list[str]:
    parts = [x.strip() for x in str(pattern or "").split(",")]
    return [x for x in parts if x]


def _depth_for(path: Path, root: Path) -> int:
    rel = path.relative_to(root)
    # file depth by directories:
    # 2020/a.xlsx -> 1 ; 2020/old/b.csv -> 2
    return max(len(rel.parts) - 1, 0)


def _scan_files(root_path: Path, *, recursive: bool, max_depth: int) -> list[Path]:
    if not root_path.exists() or not root_path.is_dir():
        return []
    if not recursive:
        return sorted([p for p in root_path.iterdir() if p.is_file()], key=lambda p: p.as_posix())

    out: list[Path] = []
    for p in root_path.rglob("*"):
        if not p.is_file():
            continue
        if max_depth > 0 and _depth_for(p, root_path) > max_depth:
            continue
        out.append(p)
    return sorted(out, key=lambda p: p.as_posix())


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _fallback_sha(path: Path, *, file_size: int, file_mtime: datetime) -> str:
    base = f"{path.as_posix()}|{file_size}|{file_mtime.isoformat()}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def _match_pattern(file_name: str, patterns: list[str]) -> bool:
    return any(fnmatch.fnmatch(file_name.lower(), p.lower()) for p in patterns)


def _normalize_row_payload(row: dict[str, Any], *, source_key: str, dataset_version: str, file_sha256: str, row_index: int) -> dict[str, Any]:
    regno_raw = None
    for k, v in row.items():
        key = str(k or "").strip()
        if not key:
            continue
        if any(h in key for h in REGNO_HINT_KEYS):
            txt = str(v).strip() if v is not None else ""
            if txt:
                regno_raw = txt
                break

    regno_norm = normalize_registration_no(regno_raw) if regno_raw else None
    if not regno_raw:
        parse_level = "FAIL"
        parse_confidence = 0.0
        parse_reason = "REGNO_MISSING"
        parse_ok = False
        regno_type = "unknown"
        regno_origin_type = "unknown"
        issuer_alias = None
        action_suffix = None
        legacy_seq = None
        is_legacy_format = False
    elif not regno_norm:
        parse_level = "FAIL"
        parse_confidence = 0.0
        parse_reason = "REGNO_NORMALIZE_FAILED"
        parse_ok = False
        regno_type = "unknown"
        regno_origin_type = "unknown"
        issuer_alias = None
        action_suffix = None
        legacy_seq = None
        is_legacy_format = False
    else:
        parsed = parse_registration_no(regno_norm)
        parse_level = parsed.parse_level
        parse_confidence = float(parsed.parse_confidence)
        parse_reason = parsed.parse_reason
        parse_ok = bool(parsed.parse_ok)
        regno_type = parsed.regno_type
        regno_origin_type = parsed.origin_type
        issuer_alias = parsed.issuer_alias
        action_suffix = parsed.action_suffix
        legacy_seq = parsed.legacy_seq
        is_legacy_format = bool(parsed.is_legacy_format)
    params_non_empty = sum(1 for _, v in row.items() if v is not None and str(v).strip() != "")
    return {
        "source_key": source_key,
        "dataset_version": dataset_version,
        "file_sha256": file_sha256,
        "row_index": row_index,
        "registration_no_raw": regno_raw,
        "registration_no_norm": regno_norm,
        "regno_parse_ok": parse_ok,
        "regno_type": regno_type,
        "regno_origin_type": regno_origin_type,
        "regno_parse_level": parse_level,
        "regno_parse_confidence": parse_confidence,
        "regno_parse_reason": parse_reason,
        "issuer_alias": issuer_alias,
        "action_suffix": action_suffix,
        "legacy_seq": legacy_seq,
        "is_legacy_format": is_legacy_format,
        "params_non_empty_count": params_non_empty,
        "data": row,
    }


def _reason_code_from_payload(payload: dict[str, Any]) -> str:
    if not payload.get("registration_no_raw"):
        return "REGNO_MISSING"
    if not payload.get("registration_no_norm"):
        return "REGNO_NORMALIZE_FAILED"
    if not bool(payload.get("regno_parse_ok")):
        return "REGNO_PARSE_FAILED"
    return "OK"


def _counter_dict(counter: Counter[str], *, top_n: int | None = None) -> dict[str, int]:
    if top_n is None:
        items = counter.items()
    else:
        items = counter.most_common(top_n)
    return {str(k): int(v) for k, v in items}


def _first_non_empty_text(row: dict[str, Any], keys: tuple[str, ...]) -> tuple[str | None, str | None]:
    for key in keys:
        if key not in row:
            continue
        txt = str(row.get(key) or "").strip()
        if txt:
            return key, txt
    return None, None


def _extract_country_or_region(row: dict[str, Any]) -> CountryExtractionResult | None:
    source_field, raw_value = _first_non_empty_text(row, PRODUCTION_ADDR_KEYS)
    if not raw_value:
        source_field, raw_value = _first_non_empty_text(row, REGISTRANT_ADDR_KEYS)
    if not raw_value or not source_field:
        return None

    for pattern, norm in _REGION_PATTERNS:
        if re.search(pattern, raw_value, flags=re.IGNORECASE):
            return CountryExtractionResult(
                country_or_region=norm,
                geo_type="region",
                source_field=source_field,
                raw_value=raw_value,
                confidence=0.95,
            )
    for pattern, norm in _COUNTRY_PATTERNS:
        if re.search(pattern, raw_value, flags=re.IGNORECASE):
            return CountryExtractionResult(
                country_or_region=norm,
                geo_type="country",
                source_field=source_field,
                raw_value=raw_value,
                confidence=0.92,
            )
    return None


def _origin_bucket_from_payload(payload: dict[str, Any]) -> str:
    origin = str(payload.get("regno_origin_type") or "").strip().lower()
    if origin in {"domestic", "import", "permit", "filing"}:
        return origin
    if not payload.get("registration_no_norm"):
        return "unknown"
    p = parse_registration_no(str(payload.get("registration_no_norm") or ""))
    if p.origin_type in {"domestic", "import", "permit", "filing"}:
        return p.origin_type
    return "unknown"


def _insert_product_param(
    db: Session,
    *,
    registry_no: str,
    param_code: str,
    value_text: str,
    evidence_text: str,
    raw_document_id: str,
    confidence: float,
    extract_version: str,
    evidence_json: dict[str, Any],
) -> bool:
    res = db.execute(
        text(
            """
            INSERT INTO product_params (
                di, registry_no, product_id, param_code, value_num, value_text, unit,
                range_low, range_high, conditions, evidence_json, evidence_text, evidence_page,
                raw_document_id, confidence, extract_version, observed_at
            ) VALUES (
                NULL, :registry_no, NULL, :param_code, NULL, :value_text, NULL,
                NULL, NULL, NULL, CAST(:evidence_json AS jsonb), :evidence_text, NULL,
                CAST(:raw_document_id AS uuid), :confidence, :extract_version, NOW()
            )
            """
        ),
        {
            "registry_no": registry_no,
            "param_code": param_code,
            "value_text": value_text,
            "evidence_text": evidence_text,
            "raw_document_id": raw_document_id,
            "confidence": confidence,
            "extract_version": extract_version,
            "evidence_json": json.dumps(evidence_json, ensure_ascii=False),
        },
    )
    return bool(res.rowcount and int(res.rowcount) > 0)


def _payload_hash(payload: dict[str, Any]) -> str:
    blob = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _iter_file_rows(path: Path) -> tuple[int, int, list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    rows: list[dict[str, Any]] = []
    failed = 0

    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
        return len(rows), failed, rows

    if suffix == ".json":
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, list):
                rows = [x for x in data if isinstance(x, dict)]
                failed = max(0, len(data) - len(rows))
            elif isinstance(data, dict):
                rows = [data]
            else:
                failed = 1
        except Exception:
            failed = 1
        return len(rows), failed, rows

    if suffix == ".ndjson":
        for ln in path.read_text(encoding="utf-8").splitlines():
            s = ln.strip()
            if not s:
                continue
            try:
                obj = json.loads(s)
                if isinstance(obj, dict):
                    rows.append(obj)
                else:
                    failed += 1
            except Exception:
                failed += 1
        return len(rows), failed, rows

    if suffix in (".xlsx", ".xls"):
        # Keep dependency-light; if openpyxl is unavailable, mark as failed row.
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            for ws in wb.worksheets:
                header: list[str] | None = None
                for values in ws.iter_rows(values_only=True):
                    arr = list(values)
                    if header is None:
                        header = [str(x).strip() if x is not None else "" for x in arr]
                        continue
                    if not any(x is not None and str(x).strip() for x in arr):
                        continue
                    row: dict[str, Any] = {}
                    for idx, key in enumerate(header):
                        k = key or f"col_{idx+1}"
                        row[k] = arr[idx] if idx < len(arr) else None
                    rows.append(row)
            return len(rows), failed, rows
        except Exception:
            return 0, 1, []

    return 0, 1, []


def _create_dataset(
    db: Session,
    *,
    source_key: str,
    dataset_version: str,
    root_path: str,
    recursive: bool,
    max_depth: int,
    pattern: str,
    dry_run: bool,
) -> str:
    row = db.execute(
        text(
            """
            INSERT INTO offline_datasets (
                source_key, dataset_version, root_path, recursive, max_depth, pattern,
                files_scanned, files_imported, files_skipped, rows_written, rows_failed,
                started_at, created_at, dry_run, summary_json
            ) VALUES (
                :source_key, :dataset_version, :root_path, :recursive, :max_depth, :pattern,
                0, 0, 0, 0, 0,
                NOW(), NOW(), :dry_run, CAST(:summary_json AS jsonb)
            )
            RETURNING id::text
            """
        ),
        {
            "source_key": source_key,
            "dataset_version": dataset_version,
            "root_path": root_path,
            "recursive": recursive,
            "max_depth": max_depth,
            "pattern": pattern,
            "dry_run": dry_run,
            "summary_json": json.dumps({}, ensure_ascii=False),
        },
    ).first()
    assert row and row[0]
    return str(row[0])


def _update_dataset_finish(
    db: Session,
    *,
    dataset_id: str,
    files_scanned: int,
    files_imported: int,
    files_skipped: int,
    rows_written: int,
    rows_failed: int,
    summary_json: dict[str, Any],
) -> None:
    db.execute(
        text(
            """
            UPDATE offline_datasets
            SET files_scanned = :files_scanned,
                files_imported = :files_imported,
                files_skipped = :files_skipped,
                rows_written = :rows_written,
                rows_failed = :rows_failed,
                summary_json = CAST(:summary_json AS jsonb),
                finished_at = NOW()
            WHERE id = CAST(:dataset_id AS uuid)
            """
        ),
        {
            "dataset_id": dataset_id,
            "files_scanned": files_scanned,
            "files_imported": files_imported,
            "files_skipped": files_skipped,
            "rows_written": rows_written,
            "rows_failed": rows_failed,
            "summary_json": json.dumps(summary_json, ensure_ascii=False),
        },
    )


def _insert_dataset_file(db: Session, *, dataset_id: str, source_key: str, file: FileImportResult) -> None:
    db.execute(
        text(
            """
            INSERT INTO offline_dataset_files (
                dataset_id, source_key, storage_uri, relative_path, file_sha256,
                file_size, file_mtime, imported, skipped_reason, rows_scanned, rows_written, rows_failed
            ) VALUES (
                CAST(:dataset_id AS uuid), :source_key, :storage_uri, :relative_path, :file_sha256,
                :file_size, :file_mtime, :imported, :skipped_reason, :rows_scanned, :rows_written, :rows_failed
            )
            """
        ),
        {
            "dataset_id": dataset_id,
            "source_key": source_key,
            "storage_uri": file.storage_uri,
            "relative_path": file.relative_path,
            "file_sha256": file.file_sha256,
            "file_size": file.file_size,
            "file_mtime": file.file_mtime,
            "imported": file.imported,
            "skipped_reason": file.skipped_reason,
            "rows_scanned": file.rows_scanned,
            "rows_written": file.rows_written,
            "rows_failed": file.rows_failed,
        },
    )


def _file_already_imported(db: Session, *, source_key: str, file_sha256: str) -> bool:
    val = db.execute(
        text(
            """
            SELECT 1
            FROM offline_dataset_files
            WHERE source_key = :source_key
              AND file_sha256 = :file_sha256
              AND imported = TRUE
            LIMIT 1
            """
        ),
        {"source_key": source_key, "file_sha256": file_sha256},
    ).first()
    return bool(val)


def _insert_raw_document(db: Session, *, source_key: str, dataset_version: str, file: FileImportResult) -> str:
    row = db.execute(
        text(
            """
            INSERT INTO raw_documents (
                source, source_url, doc_type, storage_uri, sha256, fetched_at, run_id, parse_status, parse_log, error
            ) VALUES (
                :source, NULL, :doc_type, :storage_uri, :sha256, NOW(), :run_id, 'parsed', CAST(:parse_log AS jsonb), NULL
            )
            ON CONFLICT (source, run_id, sha256) DO UPDATE
            SET parse_log = raw_documents.parse_log
            RETURNING id::text
            """
        ),
        {
            "source": source_key,
            "doc_type": (Path(file.relative_path).suffix.lower().lstrip(".") or None),
            "storage_uri": file.storage_uri,
            "sha256": file.file_sha256,
            "run_id": f"dataset:{dataset_version}",
            "parse_log": json.dumps(
                {
                    "dataset_version": dataset_version,
                    "relative_path": file.relative_path,
                    "file_size": file.file_size,
                    "file_mtime": file.file_mtime.isoformat(),
                },
                ensure_ascii=False,
            ),
        },
    ).first()
    assert row and row[0]
    return str(row[0])


def _insert_raw_source_row(
    db: Session,
    *,
    source_key: str,
    payload: dict[str, Any],
    reason_code: str,
) -> bool:
    res = db.execute(
        text(
            """
            INSERT INTO raw_source_records (
                source, source_run_id, source_url, payload_hash, evidence_grade, observed_at,
                payload, parse_status, parse_error
            ) VALUES (
                :source, NULL, NULL, :payload_hash, 'C', NOW(),
                CAST(:payload AS jsonb), :parse_status, :parse_error
            )
            ON CONFLICT DO NOTHING
            """
        ),
        {
            "source": source_key,
            "payload_hash": _payload_hash(payload),
            "payload": json.dumps(payload, ensure_ascii=False),
            "parse_status": ("parsed" if reason_code == "OK" else "error"),
            "parse_error": (None if reason_code == "OK" else reason_code),
        },
    )
    return bool(res.rowcount and int(res.rowcount) > 0)


def run_source_import_files(
    db: Session,
    *,
    source_key: str,
    root_path: Path,
    recursive: bool = True,
    max_depth: int = 0,
    pattern: str = DEFAULT_PATTERN,
    only_new: bool = True,
    dry_run: bool = True,
    dataset_version: str | None = None,
) -> OfflineImportResult:
    if not str(root_path).startswith("/data/import"):
        raise ValueError("root_path must be under /data/import")

    dataset_version = str(dataset_version or _dataset_version())
    patterns = _parse_patterns(pattern or DEFAULT_PATTERN)
    root = root_path
    files = _scan_files(root, recursive=recursive, max_depth=max_depth)
    dataset_id = _create_dataset(
        db,
        source_key=source_key,
        dataset_version=dataset_version,
        root_path=str(root),
        recursive=recursive,
        max_depth=max_depth,
        pattern=",".join(patterns),
        dry_run=dry_run,
    )

    files_scanned = 0
    files_imported = 0
    files_skipped = 0
    rows_written = 0
    rows_failed = 0
    new_files_count = 0
    dup_files_count = 0
    ext_filtered_count = 0
    parse_level_counter: Counter[str] = Counter()
    parse_reason_counter: Counter[str] = Counter()
    action_suffix_counter: Counter[str] = Counter()
    issuer_alias_counter: Counter[str] = Counter()
    country_region_counter: Counter[str] = Counter()
    origin_bucket_counter: Counter[str] = Counter()
    product_params_written = 0

    for fp in files:
        st = fp.stat()
        rel = fp.relative_to(root).as_posix()
        mt = datetime.fromtimestamp(st.st_mtime, tz=timezone.utc)
        storage_uri = f"file://{fp.as_posix()}"
        files_scanned += 1

        if not _match_pattern(fp.name, patterns):
            sha = _fallback_sha(fp, file_size=int(st.st_size), file_mtime=mt)
            ext_filtered_count += 1
            files_skipped += 1
            _insert_dataset_file(
                db,
                dataset_id=dataset_id,
                source_key=source_key,
                file=FileImportResult(
                    storage_uri=storage_uri,
                    relative_path=rel,
                    file_sha256=sha,
                    file_size=int(st.st_size),
                    file_mtime=mt,
                    imported=False,
                    skipped_reason="EXT_NOT_ALLOWED",
                    rows_scanned=0,
                    rows_written=0,
                    rows_failed=0,
                ),
            )
            continue

        try:
            sha = _file_sha256(fp)
        except Exception:
            sha = _fallback_sha(fp, file_size=int(st.st_size), file_mtime=mt)
            files_skipped += 1
            _insert_dataset_file(
                db,
                dataset_id=dataset_id,
                source_key=source_key,
                file=FileImportResult(
                    storage_uri=storage_uri,
                    relative_path=rel,
                    file_sha256=sha,
                    file_size=int(st.st_size),
                    file_mtime=mt,
                    imported=False,
                    skipped_reason="FILE_READ_ERROR",
                    rows_scanned=0,
                    rows_written=0,
                    rows_failed=1,
                ),
            )
            rows_failed += 1
            continue

        if only_new and _file_already_imported(db, source_key=source_key, file_sha256=sha):
            dup_files_count += 1
            files_skipped += 1
            _insert_dataset_file(
                db,
                dataset_id=dataset_id,
                source_key=source_key,
                file=FileImportResult(
                    storage_uri=storage_uri,
                    relative_path=rel,
                    file_sha256=sha,
                    file_size=int(st.st_size),
                    file_mtime=mt,
                    imported=False,
                    skipped_reason="DUP_SHA256",
                    rows_scanned=0,
                    rows_written=0,
                    rows_failed=0,
                ),
            )
            continue

        new_files_count += 1
        file_rows_scanned, file_rows_failed, rows = _iter_file_rows(fp)
        file_rows_written = 0
        file_params_written = 0
        imported = not dry_run
        skipped_reason = None if imported else "DRY_RUN"
        raw_document_id: str | None = None
        if imported:
            raw_document_id = _insert_raw_document(
                db,
                source_key=source_key,
                dataset_version=dataset_version,
                file=FileImportResult(
                    storage_uri=storage_uri,
                    relative_path=rel,
                    file_sha256=sha,
                    file_size=int(st.st_size),
                    file_mtime=mt,
                    imported=True,
                    skipped_reason=None,
                    rows_scanned=file_rows_scanned,
                    rows_written=0,
                    rows_failed=file_rows_failed,
                ),
            )
        file_param_seen: set[tuple[str, str, str]] = set()
        for idx, row in enumerate(rows, start=1):
            payload = _normalize_row_payload(
                row,
                source_key=source_key,
                dataset_version=dataset_version,
                file_sha256=sha,
                row_index=idx,
            )
            parse_level_counter[str(payload.get("regno_parse_level") or "FAIL")] += 1
            parse_reason_counter[str(payload.get("regno_parse_reason") or "UNKNOWN_PATTERN")] += 1
            if payload.get("action_suffix"):
                action_suffix_counter[str(payload["action_suffix"])] += 1
            if payload.get("issuer_alias"):
                issuer_alias_counter[str(payload["issuer_alias"])] += 1
            country_info = _extract_country_or_region(payload.get("data") if isinstance(payload.get("data"), dict) else {})
            if country_info is not None:
                country_region_counter[country_info.country_or_region] += 1
            origin_bucket = _origin_bucket_from_payload(payload)
            origin_bucket_counter[origin_bucket] += 1
            if imported:
                reason_code = _reason_code_from_payload(payload)
                if _insert_raw_source_row(db, source_key=source_key, payload=payload, reason_code=reason_code):
                    file_rows_written += 1
                regno_norm = str(payload.get("registration_no_norm") or "").strip()
                if regno_norm and raw_document_id:
                    # Keep params de-duplicated inside one file/doc to avoid exploding repeated address rows.
                    pkey = (regno_norm, "origin_bucket", origin_bucket)
                    if pkey not in file_param_seen:
                        if _insert_product_param(
                            db,
                            registry_no=regno_norm,
                            param_code="origin_bucket",
                            value_text=origin_bucket,
                            evidence_text=f"derived from registration_no ({payload.get('regno_parse_level')})",
                            raw_document_id=raw_document_id,
                            confidence=(0.9 if origin_bucket != "unknown" else 0.5),
                            extract_version=ORIGIN_PARAM_EXTRACT_VERSION,
                            evidence_json={
                                "dataset_version": dataset_version,
                                "parse_level": payload.get("regno_parse_level"),
                                "parse_reason": payload.get("regno_parse_reason"),
                            },
                        ):
                            file_params_written += 1
                        file_param_seen.add(pkey)
                    if country_info is not None:
                        pkey = (regno_norm, "country_or_region", country_info.country_or_region)
                        if pkey not in file_param_seen:
                            if _insert_product_param(
                                db,
                                registry_no=regno_norm,
                                param_code="country_or_region",
                                value_text=country_info.country_or_region,
                                evidence_text=f"derived from {country_info.source_field}",
                                raw_document_id=raw_document_id,
                                confidence=country_info.confidence,
                                extract_version=COUNTRY_PARAM_EXTRACT_VERSION,
                                evidence_json={
                                    "dataset_version": dataset_version,
                                    "geo_type": country_info.geo_type,
                                    "source_field": country_info.source_field,
                                    "raw_value": country_info.raw_value,
                                    "display_label": "国家/地区",
                                },
                            ):
                                file_params_written += 1
                            file_param_seen.add(pkey)

        rows_written += file_rows_written
        product_params_written += file_params_written
        rows_failed += file_rows_failed
        if imported:
            files_imported += 1
        else:
            files_skipped += 1
        _insert_dataset_file(
            db,
            dataset_id=dataset_id,
            source_key=source_key,
            file=FileImportResult(
                storage_uri=storage_uri,
                relative_path=rel,
                file_sha256=sha,
                file_size=int(st.st_size),
                file_mtime=mt,
                imported=imported,
                skipped_reason=skipped_reason,
                rows_scanned=file_rows_scanned,
                rows_written=file_rows_written,
                rows_failed=file_rows_failed,
            ),
        )

    summary_json = {
        "parse_level_distribution": _counter_dict(parse_level_counter),
        "top_parse_reasons": [{"reason": k, "count": v} for k, v in _counter_dict(parse_reason_counter, top_n=5).items()],
        "action_suffix_counts": _counter_dict(action_suffix_counter),
        "issuer_alias_counts": _counter_dict(issuer_alias_counter, top_n=5),
        "country_region_counts": _counter_dict(country_region_counter, top_n=20),
        "origin_bucket_counts": _counter_dict(origin_bucket_counter),
        "product_params_written": int(product_params_written),
    }
    _update_dataset_finish(
        db,
        dataset_id=dataset_id,
        files_scanned=files_scanned,
        files_imported=files_imported,
        files_skipped=files_skipped,
        rows_written=rows_written,
        rows_failed=rows_failed,
        summary_json=summary_json,
    )

    return OfflineImportResult(
        source_key=source_key,
        dataset_version=dataset_version,
        dataset_id=dataset_id,
        root_path=str(root),
        recursive=recursive,
        max_depth=max_depth,
        pattern=",".join(patterns),
        only_new=only_new,
        dry_run=dry_run,
        files_scanned=files_scanned,
        files_imported=files_imported,
        files_skipped=files_skipped,
        rows_written=rows_written,
        rows_failed=rows_failed,
        new_files_count=new_files_count,
        dup_files_count=dup_files_count,
        ext_filtered_count=ext_filtered_count,
        parse_level_distribution=_counter_dict(parse_level_counter),
        top_parse_reasons=[{"reason": k, "count": v} for k, v in _counter_dict(parse_reason_counter, top_n=5).items()],
        action_suffix_counts=_counter_dict(action_suffix_counter),
        issuer_alias_counts=_counter_dict(issuer_alias_counter, top_n=5),
        country_region_counts=_counter_dict(country_region_counter, top_n=20),
        origin_bucket_counts=_counter_dict(origin_bucket_counter),
        product_params_written=product_params_written,
    )
